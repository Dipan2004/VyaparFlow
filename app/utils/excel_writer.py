"""
utils/excel_writer.py
---------------------
Excel persistence utility for Notiflow.  (FIX 4 — config-driven paths)

All file paths come from app/config.py — no hardcoded paths in this module.

Single-file, multi-sheet store:  DATA_FILE  (default: data/notiflow_data.xlsx)

Sheets and their column schemas:
    Orders     — order records
    Ledger     — payment and credit entries
    Returns    — return / exchange requests
    Inventory  — stock movement delta log
    Invoices   — generated invoice records

Public API
----------
append_row(sheet_name, record)          — append one row, atomic save
append_rows(sheet_name, records)        — append many rows, one save
read_sheet(sheet_name) -> DataFrame     — read sheet into pandas DataFrame
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd

from app.config import DATA_FILE   # FIX 4: single source of truth

logger = logging.getLogger(__name__)

EXCEL_FILE = Path(DATA_FILE)        # re-export for modules that imported it directly

# ---------------------------------------------------------------------------
# Canonical column schemas
# ---------------------------------------------------------------------------

SHEET_SCHEMAS: dict[str, list[str]] = {
    "Orders": [
        "order_id", "timestamp", "customer", "item", "quantity", "status"
    ],
    "Ledger": [
        "entry_id", "timestamp", "type", "customer", "item",
        "quantity", "amount", "payment_type", "status"
    ],
    "Returns": [
        "return_id", "timestamp", "customer", "item", "reason", "status"
    ],
    "Inventory": [
        "timestamp", "item", "change", "direction", "reference_id", "note"
    ],
    "Invoices": [
        "invoice_id", "timestamp", "order_id", "customer",
        "item", "quantity", "unit_price", "total_amount", "status"
    ],
}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _ensure_file() -> None:
    """Create the Excel file with all sheets if it does not exist."""
    from openpyxl import Workbook  # lazy import — not needed at module load time
    if EXCEL_FILE.exists():
        return

    EXCEL_FILE.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Creating new Excel file: %s", EXCEL_FILE)
    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, columns in SHEET_SCHEMAS.items():
        ws = wb.create_sheet(title=sheet_name)
        ws.append(columns)

    wb.save(EXCEL_FILE)
    logger.info("Excel file created with sheets: %s", list(SHEET_SCHEMAS.keys()))


def _ensure_sheet(wb, sheet_name: str) -> None:
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        columns = SHEET_SCHEMAS.get(sheet_name, [])
        if columns:
            ws.append(columns)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def append_row(sheet_name: str, record: dict) -> None:
    """
    Append a single record to the named Excel sheet.

    Missing columns default to None. Extra keys in record are ignored.
    File and sheet are created automatically if they don't exist.

    Args:
        sheet_name: e.g. "Orders", "Ledger"
        record:     Dict of column → value pairs.

    Raises:
        ValueError: Unknown sheet_name.
    """
    if sheet_name not in SHEET_SCHEMAS:
        raise ValueError(
            f"Unknown sheet '{sheet_name}'. Valid: {list(SHEET_SCHEMAS)}"
        )

    from openpyxl import load_workbook  # lazy import
    _ensure_file()
    wb = load_workbook(EXCEL_FILE)
    _ensure_sheet(wb, sheet_name)

    ws      = wb[sheet_name]
    columns = SHEET_SCHEMAS[sheet_name]
    ws.append([record.get(col) for col in columns])

    wb.save(EXCEL_FILE)
    logger.debug("Row appended to '%s': %s", sheet_name, record)


def append_rows(sheet_name: str, records: list[dict]) -> None:
    """Append multiple records in one file open/save cycle."""
    if not records:
        return
    if sheet_name not in SHEET_SCHEMAS:
        raise ValueError(f"Unknown sheet '{sheet_name}'.")

    from openpyxl import load_workbook  # lazy import
    _ensure_file()
    wb = load_workbook(EXCEL_FILE)
    _ensure_sheet(wb, sheet_name)

    ws      = wb[sheet_name]
    columns = SHEET_SCHEMAS[sheet_name]
    for record in records:
        ws.append([record.get(col) for col in columns])

    wb.save(EXCEL_FILE)
    logger.debug("Appended %d rows to '%s'", len(records), sheet_name)


def read_sheet(sheet_name: str) -> pd.DataFrame:
    """
    Read a sheet into a DataFrame.

    Returns an empty DataFrame (with correct columns) if the file or
    sheet does not exist yet.
    """
    columns = SHEET_SCHEMAS.get(sheet_name, [])
    if not EXCEL_FILE.exists():
        return pd.DataFrame(columns=columns)
    try:
        return pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
    except Exception as exc:
        logger.warning("Could not read sheet '%s': %s", sheet_name, exc)
        return pd.DataFrame(columns=columns)